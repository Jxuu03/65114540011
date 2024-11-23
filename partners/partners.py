from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('Me & My partner.xlsx')
    groups = []

    for row in wb.active.iter_rows(min_row=2):
        values = [cell.value for cell in row]

        if values[1] is not None and values[2] is not None:
            pair = (values[1], values[2])

            if values[1] < values[2]:
                pair = (values[2], values[1])
            if pair not in groups:
                groups.append(pair)


    print(f"Total number of valid pairs: {len(groups)}")
    for pair in groups:
        print(pair)
    